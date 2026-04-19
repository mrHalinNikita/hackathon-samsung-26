from pathlib import Path

import openpyxl

from src.parsers.base import BaseParser, ParsedContent


class XlsxParser(BaseParser):
    @property
    def supported_extensions(self) -> list[str]:
        return [".xlsx", ".xlsm"]

    async def parse(self, filepath: Path) -> ParsedContent:
        lines: list[str] = []
        metadata = {"path": str(filepath), "sheets": [], "parser_mode": "structured"}

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows = list(sheet.iter_rows(values_only=True))
                if not rows:
                    continue
                headers = [str(c).strip() if c is not None else f"col_{i}" for i, c in enumerate(rows[0], start=1)]
                row_count = 0
                for idx, row in enumerate(rows[1:], start=1):
                    pairs = []
                    for key, val in zip(headers, row):
                        if val is None:
                            continue
                        sval = str(val).strip()
                        if not sval:
                            continue
                        pairs.append(f"{key}: {sval}")
                    if pairs:
                        row_count += 1
                        lines.append(f"SHEET[{sheet_name}] ROW[{idx}] | " + " | ".join(pairs))
                metadata["sheets"].append({"name": sheet_name, "rows": row_count, "columns": len(headers)})
        except Exception as e:
            return ParsedContent(text="", metadata=metadata, errors=[f"XLSX parse error: {e}"])

        full_text = "\n".join(lines)
        return ParsedContent(text=full_text, metadata=metadata, word_count=len(full_text.split()), char_count=len(full_text))
